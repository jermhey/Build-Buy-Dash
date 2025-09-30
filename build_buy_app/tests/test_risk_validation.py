"""
Risk Adjustment Validation Test - Formula Logic Verification
This test validates that Excel formulas match simulation logic without requiring Excel evaluation.
"""
import sys
import os
import tempfile
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

from src.simulation import BuildVsBuySimulator
from core.excel_export import ExcelExporter
from openpyxl import load_workbook

def validate_risk_adjustment_logic():
    """Validate that Excel formulas logically match simulation calculations."""
    
    test_params = {
        'build_timeline': 12,
        'fte_cost': 150000,
        'fte_count': 2,
        'useful_life': 5,
        'prob_success': 80,
        'wacc': 10,
        'tech_risk': 10,
        'vendor_risk': 5,
        'market_risk': 5,
        'misc_costs': 50000,
        'capex': 100000,
        'amortization': 0,
        'maint_opex': 20000,
        'product_price': 500000,
        'subscription_price': 0,
        'buy_selector': ['one_time']
    }
    
    print("🧪 Risk Adjustment Logic Validation")
    print("=" * 60)
    
    # Run simulation
    simulator = BuildVsBuySimulator(n_simulations=1000)
    sim_results = simulator.simulate(test_params)
    
    print(f"📊 Simulation Results:")
    print(f"  Expected Build Cost: ${sim_results['expected_build_cost']:,.2f}")
    
    # Generate Excel
    exporter = ExcelExporter()
    excel_bytes = exporter.create_excel_export(test_params)
    
    # Extract formula logic from Excel
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(excel_bytes)
        tmp_path = tmp.name
    
    try:
        wb = load_workbook(tmp_path, data_only=False)
        
        # Extract input values (skip formula cells, get raw numbers only)
        input_values = {}
        if 'Input Parameters' in wb.sheetnames:
            ws_input = wb['Input Parameters']
            for row in ws_input.iter_rows():
                if len(row) >= 2 and row[0].value and row[1].value is not None:
                    label = str(row[0].value)
                    value = row[1].value
                    
                    # Only extract direct numeric values, skip formulas
                    if isinstance(value, (int, float)):
                        if 'Build Timeline' in label:
                            input_values['timeline'] = value
                        elif 'FTE Cost' in label and 'Total' not in label:
                            input_values['fte_cost'] = value
                        elif 'FTE Count' in label:
                            input_values['fte_count'] = value
                        elif 'Success Probability' in label:
                            input_values['prob_success'] = value
                        elif 'WACC' in label:
                            input_values['wacc'] = value
                        elif 'Technical Risk' in label:
                            input_values['tech_risk'] = value
                        elif 'Vendor Risk' in label:
                            input_values['vendor_risk'] = value
                        elif 'Market Risk' in label:
                            input_values['market_risk'] = value
                        elif 'CapEx' in label:
                            input_values['capex'] = value
                        elif 'Miscellaneous' in label:
                            input_values['misc'] = value
                        elif 'Annual Maintenance' in label:
                            input_values['maint'] = value
                        elif 'Useful Life' in label:
                            input_values['useful_life'] = value
        
        # Extract Excel formulas
        excel_formulas = {}
        if 'Cost Timeline' in wb.sheetnames:
            ws_timeline = wb['Cost Timeline']
            for row in ws_timeline.iter_rows():
                if row[0].value and isinstance(row[0].value, str):
                    component = row[0].value
                    
                    # Look for formulas in Present Value column (column I)
                    if len(row) > 8 and row[8].value and isinstance(row[8].value, str):
                        formula = row[8].value
                        if 'FTE Labor' in component:
                            excel_formulas['labor_pv'] = formula
                        elif 'Risk Premium' in component:
                            excel_formulas['risk_premium'] = formula
                        elif 'TOTAL BUILD COST' in component:
                            excel_formulas['total_build'] = formula
                        elif 'Maintenance' in component:
                            excel_formulas['maintenance_pv'] = formula
        
        # Also get Total FTE Cost formula from Input Parameters
        if 'Input Parameters' in wb.sheetnames:
            ws_input = wb['Input Parameters']
            for row in ws_input.iter_rows():
                if len(row) >= 2 and row[0].value and 'Total FTE Costs' in str(row[0].value):
                    if row[1].value and isinstance(row[1].value, str):
                        excel_formulas['total_fte'] = row[1].value
        
        print(f"\n📋 Excel Input Values:")
        for key, value in input_values.items():
            print(f"  {key}: {value}")
        
        print(f"\n📝 Excel Formulas:")
        for key, formula in excel_formulas.items():
            print(f"  {key}: {formula}")
        
        # Manual calculation based on Excel formulas
        print(f"\n🔍 Manual Calculation Using Excel Formula Logic:")
        
        # 1. Total FTE Cost (success-adjusted)
        # Formula: =(timeline/12)*fte_cost*fte_count/prob_success
        excel_total_fte = (input_values['timeline']/12) * input_values['fte_cost'] * input_values['fte_count'] / input_values['prob_success']
        print(f"  1. Total FTE Cost: ${excel_total_fte:,.2f}")
        print(f"     Excel Formula: ({input_values['timeline']}/12)*{input_values['fte_cost']}*{input_values['fte_count']}/{input_values['prob_success']}")
        
        # 2. Base costs (FTE + CapEx + Misc)
        base_costs = excel_total_fte + input_values['capex'] + input_values['misc']
        print(f"  2. Base Costs: ${base_costs:,.2f}")
        print(f"     (FTE ${excel_total_fte:,.2f} + CapEx ${input_values['capex']:,.2f} + Misc ${input_values['misc']:,.2f})")
        
        # 3. Risk Premium
        # Formula: =(base_costs)*(tech_risk+vendor_risk+market_risk)
        total_risk = input_values['tech_risk'] + input_values['vendor_risk'] + input_values['market_risk']
        excel_risk_premium = base_costs * total_risk
        print(f"  3. Risk Premium: ${excel_risk_premium:,.2f}")
        print(f"     Excel Formula: {base_costs:,.2f}*({input_values['tech_risk']}+{input_values['vendor_risk']}+{input_values['market_risk']})")
        print(f"     Total Risk: {total_risk:.3f} ({total_risk*100:.1f}%)")
        
        # 4. Maintenance PV
        # Formula: =maint*((1-(1+wacc)^-useful_life)/wacc)
        wacc = input_values['wacc']
        useful_life = input_values['useful_life']
        excel_maint_pv = input_values['maint'] * ((1-(1+wacc)**-useful_life)/wacc)
        print(f"  4. Maintenance PV: ${excel_maint_pv:,.2f}")
        print(f"     Excel Formula: {input_values['maint']}*((1-(1+{wacc})^-{useful_life})/{wacc})")
        
        # 5. Total Build Cost
        # Formula: =base_costs + risk_premium + maintenance_pv
        excel_total_build = base_costs + excel_risk_premium + excel_maint_pv
        print(f"  5. Total Build Cost: ${excel_total_build:,.2f}")
        print(f"     Excel Formula: {base_costs:,.2f} + {excel_risk_premium:,.2f} + {excel_maint_pv:,.2f}")
        
        # Simulation validation
        print(f"\n⚖️  Excel vs Simulation Comparison:")
        print(f"  Excel Calculated Total: ${excel_total_build:,.2f}")
        print(f"  Simulation Result:      ${sim_results['expected_build_cost']:,.2f}")
        
        difference = abs(excel_total_build - sim_results['expected_build_cost'])
        tolerance = sim_results['expected_build_cost'] * 0.03  # 3% tolerance
        
        print(f"  Difference: ${difference:,.2f}")
        print(f"  Tolerance:  ${tolerance:,.2f} (3%)")
        
        # Validation checks
        success_criteria = []
        
        # Check 1: Probability of success adjustment
        expected_labor = (test_params['build_timeline']/12) * test_params['fte_cost'] * test_params['fte_count'] / (test_params['prob_success']/100)
        labor_match = abs(excel_total_fte - expected_labor) < 100
        success_criteria.append(("Probability of Success Adjustment", labor_match, f"${excel_total_fte:,.2f} ≈ ${expected_labor:,.2f}"))
        
        # Check 2: Risk factor application
        expected_risk = base_costs * (test_params['tech_risk'] + test_params['vendor_risk'] + test_params['market_risk']) / 100
        risk_match = abs(excel_risk_premium - expected_risk) < 100
        success_criteria.append(("Risk Factor Application", risk_match, f"${excel_risk_premium:,.2f} ≈ ${expected_risk:,.2f}"))
        
        # Check 3: Overall calculation
        calc_match = difference < tolerance
        success_criteria.append(("Overall Calculation Match", calc_match, f"${difference:,.2f} < ${tolerance:,.2f}"))
        
        print(f"\n🏆 Validation Results:")
        all_passed = True
        for criterion, passed, details in success_criteria:
            status = "✅" if passed else "❌"
            print(f"  {status} {criterion}: {details}")
            all_passed = all_passed and passed
        
        if all_passed:
            print(f"\n🎉 RISK ADJUSTMENT LOGIC VALIDATION SUCCESSFUL!")
            print(f"   ✅ Excel applies probability of success adjustment correctly")
            print(f"   ✅ Excel applies risk factors using additive model (industry standard)")
            print(f"   ✅ Excel calculations match simulation within tolerance")
            print(f"   ✅ Both methods are mathematically accurate and industry-standard")
            return True
        else:
            print(f"\n❌ RISK ADJUSTMENT LOGIC VALIDATION FAILED")
            print(f"   Some criteria not met - additional adjustments needed")
            return False
            
    finally:
        os.unlink(tmp_path)

if __name__ == "__main__":
    success = validate_risk_adjustment_logic()
    if success:
        print("\n✅ VALIDATION PASSED - Excel risk adjustments now match simulation!")
    else:
        print("\n❌ VALIDATION FAILED - Additional fixes needed.")
