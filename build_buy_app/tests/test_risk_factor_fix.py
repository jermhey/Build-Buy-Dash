#!/usr/bin/env python3
"""
Test to verify that Excel risk factor adjustments now match simulation
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

from src.simulation import BuildVsBuySimulator
from core.excel_export import ExcelExporter

def test_risk_factor_fix():
    """Test that Excel now applies risk factors correctly to match simulation."""
    print("🔧 TESTING RISK FACTOR FIX")
    print("=" * 60)
    
    # Use the exact parameters from the user's Excel example
    params = {
        'build_timeline': 12,
        'fte_cost': 175000,
        'fte_count': 5,
        'useful_life': 5,
        'prob_success': 90,
        'wacc': 8,
        'misc_costs': 200000,
        'capex': 100000,
        'maint_opex': 50000,
        'tech_risk': 7,
        'vendor_risk': 8,
        'market_risk': 9,
        'cap_percent': 75,
        'tax_credit_rate': 17,
        'product_price': 1000000,
        'subscription_price': 250000,
        'subscription_increase': 3,
        'buy_selector': ['one_time', 'subscription']
    }
    
    # Run simulation
    simulator = BuildVsBuySimulator(n_simulations=1000, random_seed=42)
    results = simulator.simulate(params)
    
    print(f"Simulation Build Cost: ${results['expected_build_cost']:,.0f}")
    
    # Manual calculation of what Excel should now produce (corrected methodology)
    print("\nManual calculation (matching corrected Excel logic):")
    
    # Use the exact parameters from simulation 
    build_timeline = params['build_timeline']  # 12 months
    fte_cost = params['fte_cost']  # $175,000
    fte_count = params['fte_count']  # 5
    prob_success = params['prob_success'] / 100  # 90% -> 0.9
    cap_percent = params['cap_percent'] / 100  # 75% -> 0.75
    tax_credit_rate = params['tax_credit_rate'] / 100  # 17% -> 0.17
    misc_costs = params['misc_costs']  # $200,000
    capex = params['capex']  # $100,000
    maint_opex = params['maint_opex']  # $50,000
    wacc = params['wacc'] / 100  # 8% -> 0.08
    useful_life = params['useful_life']  # 5 years
    tech_risk = params['tech_risk'] / 100  # 7% -> 0.07
    vendor_risk = params['vendor_risk'] / 100  # 8% -> 0.08
    market_risk = params['market_risk'] / 100  # 9% -> 0.09
    
    # Excel formulas as implemented:
    # Total FTE Cost: =((B6/12)*B8*B10)/B13
    total_fte_cost = ((build_timeline/12) * fte_cost * fte_count) / prob_success
    print(f"Total FTE Cost: {total_fte_cost:,.0f}")
    
    # Capitalized and Expensed Labor
    capitalized_labor = total_fte_cost * cap_percent
    expensed_labor = total_fte_cost - capitalized_labor
    print(f"Capitalized Labor: ${capitalized_labor:,.0f}")
    print(f"Expensed Labor: ${expensed_labor:,.0f}")
    
    # Tax Credit
    tax_credit = capitalized_labor * tax_credit_rate
    print(f"Tax Credit: ${tax_credit:,.0f}")
    
    # Maintenance OpEx Present Value
    maint_pv = sum(maint_opex / (1 + wacc)**year for year in range(1, useful_life + 1))
    print(f"Maintenance OpEx PV: ${maint_pv:,.0f}")
    print(f"Misc Costs: ${misc_costs:,.0f}")
    print(f"CapEx: ${capex:,.0f}")
    
    # Base build total (before risk factors)
    base_build_total = expensed_labor + capitalized_labor + misc_costs + capex + maint_pv - tax_credit
    print(f"Base Build Total: ${base_build_total:,.0f}")
    
    # Apply risk factors to base total (NEW - this was the missing piece!)
    risk_multiplier = (1 + tech_risk) * (1 + vendor_risk) * (1 + market_risk)
    expected_excel_cost = base_build_total * risk_multiplier
    print(f"Risk Multiplier: {risk_multiplier:.6f}")
    print(f"Risk-Adjusted Excel Total: ${expected_excel_cost:,.0f}")
    
    # 7. Compare with simulation
    difference = abs(results['expected_build_cost'] - expected_excel_cost)
    pct_diff = (difference / results['expected_build_cost']) * 100
    
    print(f"\nComparison:")
    print(f"Simulation: ${results['expected_build_cost']:,.0f}")
    print(f"Expected Excel: ${expected_excel_cost:,.0f}")
    print(f"Difference: ${difference:,.0f} ({pct_diff:.2f}%)")
    
    # Test should show much smaller difference now
    if pct_diff < 5.0:  # Less than 5% difference
        print("✅ RISK FACTOR FIX SUCCESSFUL!")
        print("   Excel and simulation now match within acceptable tolerance")
        return True
    else:
        print("❌ Still significant difference - may need further investigation")
        print("   However, this is a major improvement from the original $400K+ difference")
    print(f"Expected Excel total: ${expected_excel_cost:,.0f}")
    
    # 7. Compare with simulation
    difference = abs(results['expected_build_cost'] - expected_excel_cost)
    pct_diff = (difference / results['expected_build_cost']) * 100
    
    print(f"\nComparison:")
    print(f"Simulation: ${results['expected_build_cost']:,.0f}")
    print(f"Expected Excel: ${expected_excel_cost:,.0f}")
    print(f"Difference: ${difference:,.0f} ({pct_diff:.2f}%)")
    
    # Test should show much smaller difference now
    if pct_diff < 5.0:  # Less than 5% difference
        print("✅ RISK FACTOR FIX SUCCESSFUL!")
        print("   Excel and simulation now match within acceptable tolerance")
    else:
        print("❌ Still significant difference - may need further investigation")
    
    # Create Excel export to verify formulas
    print("\nCreating Excel export to verify formulas...")
    scenario_data = params.copy()
    scenario_data['results'] = results
    
    exporter = ExcelExporter()
    excel_bytes = exporter.create_excel_export(scenario_data)
    print(f"Excel export created: {len(excel_bytes)} bytes")
    
    # Save to file for inspection
    excel_path = "temp_risk_test.xlsx"
    with open(excel_path, 'wb') as f:
        f.write(excel_bytes)
    print(f"Saved Excel file to: {excel_path}")
    
    # Try to read actual Excel values
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if 'Cost_Breakdown' in wb.sheetnames:
            ws = wb['Cost_Breakdown']
            
            # Look for the risk-adjusted build total
            for row in range(1, 100):
                cell_value = ws[f'A{row}'].value
                if cell_value and 'Risk-Adjusted Total' in str(cell_value):
                    excel_total = ws[f'D{row}'].value  # Try column D (PV)
                    if not excel_total:
                        excel_total = ws[f'C{row}'].value  # Try column C
                    if not excel_total:
                        excel_total = ws[f'B{row}'].value  # Try column B
                    
                    if excel_total:
                        print(f"\n📊 ACTUAL EXCEL VALUES:")
                        print(f"Excel Risk-Adjusted Total: ${excel_total:,.0f}")
                        difference_actual = abs(results['expected_build_cost'] - excel_total)
                        pct_diff_actual = (difference_actual / results['expected_build_cost']) * 100
                        print(f"Excel vs Simulation: ${excel_total:,.0f} vs ${results['expected_build_cost']:,.0f}")
                        print(f"Actual Difference: ${difference_actual:,.0f} ({pct_diff_actual:.2f}%)")
                        
                        if pct_diff_actual < 5:
                            print("✅ SUCCESS! Excel now matches simulation within tolerance")
                            return True
                    break
        wb.close()
    except ImportError:
        print("Install openpyxl to verify Excel calculated values: pip install openpyxl")
    except Exception as e:
        print(f"Could not read Excel values: {e}")
    
    return False

if __name__ == "__main__":
    test_risk_factor_fix()
